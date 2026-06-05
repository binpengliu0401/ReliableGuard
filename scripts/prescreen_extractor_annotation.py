"""Deterministic pre-screen for the extractor-annotation precision sheet (Table 1).

Reduces the human's workload without acting as the judge: a rule-based (NOT model-based)
check decides whether each predicted claim's value/entity literally appears in the agent
answer. High-confidence matches are pre-filled `valid=1`; everything the rule cannot
confirm is left blank and flagged `NEEDS_REVIEW` for human adjudication. The human keeps
the final say on every label; only the `NEEDS_REVIEW` rows actually need attention.

This stays methodologically clean (no LLM grading an LLM): the pre-screen is pure string /
number matching, so it introduces no model bias into the gold labels.

Outputs (overwrites):
  - eval/annotation/extractor_annotation_claims.csv  (canonical, + `review`/`prescreen_reason`)
  - eval/annotation/extractor_annotation_claims.xlsx (NEEDS_REVIEW rows highlighted yellow)

Usage: python3 scripts/prescreen_extractor_annotation.py
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPO_ROOT = Path(__file__).resolve().parents[1]
ANN_DIR = REPO_ROOT / "eval" / "annotation"
CLAIMS_CSV = ANN_DIR / "extractor_annotation_claims.csv"
MANIFEST_CSV = ANN_DIR / "sample_manifest.csv"
CLAIMS_XLSX = ANN_DIR / "extractor_annotation_claims.xlsx"


def _norm_num(text: str) -> str | None:
    text = str(text).replace(",", "").strip()
    try:
        return ("%f" % float(text)).rstrip("0").rstrip(".")
    except ValueError:
        return None


def _in_answer(answer: str, value) -> bool:
    if value is None:
        return False
    text = str(value).strip().lower()
    if not text:
        return False
    if text in answer.lower():
        return True
    norm = _norm_num(text)
    if norm is not None:
        for token in re.findall(r"[\d][\d,.]*", answer):
            if _norm_num(token) == norm:
                return True
    return False


def _all_in_answer(answer: str, values) -> bool:
    items = [v for v in values if str(v).strip()]
    return bool(items) and all(_in_answer(answer, v) for v in items)


def prescreen(claim: dict, answer: str) -> tuple[str, str]:
    """Return (valid, reason): valid is '1' (auto-pass) or '' (needs review)."""
    ctype = claim.get("claim_type")
    value = claim.get("value")
    entities = claim.get("entities") or {}

    if ctype == "existence":
        if _all_in_answer(answer, list(entities.values())):
            return "1", "auto: existence entity present in answer"
        return "", "REVIEW: existence entity not found verbatim — read the answer"

    if value is not None and str(value).strip() != "":
        if isinstance(value, list):
            if _all_in_answer(answer, value):
                return "1", "auto: all listed values present in answer"
            return "", "REVIEW: one or more listed values not found in answer"
        if _in_answer(answer, value):
            return "1", f"auto: value '{value}' present in answer"
        return "", f"REVIEW: value '{value}' not found in answer (possible parse error)"

    return "", f"REVIEW: {ctype} claim has no checkable value — judge by reading"


def load_traces() -> dict[str, tuple[str, list[dict]]]:
    manifest = {r["sample_id"]: r["trace_file"] for r in csv.DictReader(MANIFEST_CSV.open())}
    cache: dict[str, tuple[str, list[dict]]] = {}
    for sid, rel in manifest.items():
        data = json.loads((REPO_ROOT / rel).read_text(encoding="utf-8"))
        cache[sid] = (data.get("answer") or "", [t["claim"] for t in data.get("traces", [])])
    return cache


def main() -> None:
    traces = load_traces()
    rows = list(csv.DictReader(CLAIMS_CSV.open(encoding="utf-8")))

    auto = review = 0
    for row in rows:
        answer, claims = traces[row["sample_id"]]
        idx = int(row["claim_idx"]) - 1
        claim = claims[idx] if 0 <= idx < len(claims) else {}
        valid, reason = prescreen(claim, answer)
        row["valid"] = valid
        row["review"] = "" if valid == "1" else "NEEDS_REVIEW"
        row["prescreen_reason"] = reason
        if valid == "1":
            auto += 1
        else:
            review += 1

    fieldnames = ["sample_id", "domain", "stratum", "claim_idx", "claim_text",
                  "claim_struct", "query", "answer", "review", "prescreen_reason",
                  "valid", "note"]
    with CLAIMS_CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    _write_xlsx(rows, fieldnames)

    print(f"pre-screened {len(rows)} claims: auto-filled valid=1 for {auto}, "
          f"flagged NEEDS_REVIEW for {review} ({100 * review / len(rows):.0f}%).")
    print(f"wrote: {CLAIMS_CSV}")
    print(f"       {CLAIMS_XLSX}  (NEEDS_REVIEW rows highlighted; fill the blank `valid` cells)")


def _write_xlsx(rows: list[dict], fieldnames: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "claims"
    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    wrap_top = Alignment(vertical="top", wrap_text=True)

    ws.append(fieldnames)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top")

    for row in rows:
        ws.append([row.get(k, "") for k in fieldnames])
        if row["review"] == "NEEDS_REVIEW":
            for cell in ws[ws.max_row]:
                cell.fill = yellow

    widths = {"claim_text": 48, "claim_struct": 40, "answer": 70, "query": 30,
              "prescreen_reason": 38, "review": 14, "valid": 7, "note": 24}
    for i, name in enumerate(fieldnames, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = widths.get(name, 12)
        if name in {"claim_text", "claim_struct", "answer"}:
            for cell in ws[letter][1:]:
                cell.alignment = wrap_top

    # 0/1 dropdown on the valid column
    valid_letter = get_column_letter(fieldnames.index("valid") + 1)
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    dv.add(f"{valid_letter}2:{valid_letter}{ws.max_row}")
    ws.add_data_validation(dv)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{ws.max_row}"
    wb.save(CLAIMS_XLSX)


if __name__ == "__main__":
    main()
